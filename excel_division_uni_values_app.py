from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook
from copy import copy
import tkinter as tk
from tkinter import filedialog
import requests
from io import BytesIO
from PIL import Image, ImageTk

root = tk.Tk()
root.title("Excel Division by unique values in a column")

# Create and place labels and entries
tk.Label(root, text="Path to excel:").grid(row=0, column=0, padx=10, pady=5, sticky="e")
open_name_var = tk.StringVar()
tk.Entry(root, textvariable=open_name_var, width=50).grid(row=0, column=1, padx=10, pady=5)

tk.Label(root, text="Name of worksheet:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
sheet_name_var = tk.StringVar()
tk.Entry(root, textvariable=sheet_name_var, width=50).grid(row=1, column=1, padx=10, pady=5)

tk.Label(root, text="How to name the worksheet in the divided excels:").grid(row=2, column=0, padx=10, pady=5, sticky="e")
sheet_name_new_var = tk.StringVar()
tk.Entry(root, textvariable=sheet_name_new_var, width=50).grid(row=2, column=1, padx=10, pady=5)

tk.Label(root, text="Column number (A=1, B=2, etc.):").grid(row=3, column=0, padx=10, pady=5, sticky="e")
column_num_var = tk.IntVar()
tk.Entry(root, textvariable=column_num_var, width=50).grid(row=3, column=1, padx=10, pady=5)

tk.Label(root, text="Where to save the divided excels:").grid(row=4, column=0, padx=10, pady=5, sticky="e")
save_name_var = tk.StringVar()
tk.Entry(root, textvariable=save_name_var, width=50).grid(row=4, column=1, padx=10, pady=5)

tk.Label(root, text="What do you want to name the files?").grid(row=5, column=0, padx=10, pady=5, sticky="e")
excel_name_var = tk.StringVar()
tk.Entry(root, textvariable=excel_name_var, width=50).grid(row=5, column=1, padx=10, pady=5)

# Button to open file dialog
tk.Button(root, text="Browse", command=lambda: open_name_var.set(filedialog.askopenfilename())).grid(row=0, column=2, padx=10, pady=5)
tk.Button(root, text="Browse", command=lambda:save_name_var.set(filedialog.askdirectory())).grid(row=4, column=2, padx=10, pady=5)

try:
    response = requests.get("https://raw.githubusercontent.com/MMateo1120/Excel_division_app/refs/heads/main/mz6axvogusy31.jpg")
    image_data = BytesIO(response.content)
    image = Image.open(image_data)
    resized_image = image.resize((200, int(image.height * 200 / image.width)))
    cat_image = ImageTk.PhotoImage(resized_image)
    tk.Label(root, image=cat_image).grid(row=0, column=6, rowspan=6, sticky="nwe")
except:
    pass


# Function to submit and close
def submit():
    global open_name, sheet_name, sheet_name_new, column_num, save_name, excel_name
    open_name = open_name_var.get()
    sheet_name = sheet_name_var.get()
    sheet_name_new = sheet_name_new_var.get()
    column_num = column_num_var.get()
    save_name = save_name_var.get()
    excel_name = excel_name_var.get()
    root.destroy()

# Submit button
tk.Button(root, text="Submit", command=submit).grid(row=6, column=2, padx=10, pady=10)

root.mainloop()

#Excel load
wb = load_workbook(open_name)
ws = wb[sheet_name]

#Get the header
header = []
for cell_row in ws[1]:
    header.append(cell_row.value)

# Dictionary to hold workbooks for each unique value in column 'F'
wb_dict = {}

# Iterate over the rows in the worksheet
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
    print(f"{row[0].row}.row from the total of {ws.max_row} rows")
    value = row[column_num-1].value  # Column 'F' (0-indexed as 5)

    if value not in wb_dict:
        # Create a new workbook for the unique value
        wb_dict[value] = Workbook()
        ws_new = wb_dict[value].active
        ws_new.title = sheet_name_new

        # Copy the header
        ws_new.append(header)
        for cell_orig, cell_new in zip(ws[1], ws_new[1]):
            cell_new.alignment = copy(cell_orig.alignment)
            cell_new.number_format = copy(cell_orig.number_format)
            cell_new.font = copy(cell_orig.font)
            cell_new.fill = copy(cell_orig.fill)
            cell_new.border = copy(cell_orig.border)
        
        for col_num in range(1,ws.max_column+1):
            ws_new.column_dimensions[get_column_letter(col_num)].width = ws.column_dimensions[get_column_letter(col_num)].width
            
    # Append the row to the appropriate workbook
    ws_new = wb_dict[value].active
    ws_new.append([cell.value for cell in row])

    # Copy styles
    for cell_orig, cell_new in zip(row, ws_new[ws_new.max_row]):
        cell_new.alignment = copy(cell_orig.alignment)
        cell_new.number_format = copy(cell_orig.number_format)
        cell_new.font = copy(cell_orig.font)
        cell_new.fill = copy(cell_orig.fill)
        cell_new.border = copy(cell_orig.border)

# Save all workbooks
for value, wb in wb_dict.items():
    wb.save(save_name + f"/{value}_{excel_name}.xlsx")
root.destroy()

