from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook
from copy import copy
import tkinter as tk
from tkinter import filedialog

root = tk.Tk()
root.title("Excel Division")

# Create and place labels and entries
tk.Label(root, text="Path to excel:").grid(row=0, column=0, padx=10, pady=5, sticky="e")
open_name_var = tk.StringVar()
tk.Entry(root, textvariable=open_name_var, width=50).grid(row=0, column=1, padx=10, pady=5)

tk.Label(root, text="Name of worksheet:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
sheet_name_var = tk.StringVar()
tk.Entry(root, textvariable=sheet_name_var, width=50).grid(row=1, column=1, padx=10, pady=5)

tk.Label(root, text="How to name the worksheet in the divided excels:").grid(row=2, column=0, padx=10, pady=5, sticky="e")
sheet_name_new_var = tk.StringVar()
tk.Entry(root, textvariable=sheet_name_new_var, width=50).grid(row=2, column=1, padx=10, pady=5)

tk.Label(root, text="Column number (A=1, B=2, etc.):").grid(row=3, column=0, padx=10, pady=5, sticky="e")
column_num_var = tk.IntVar()
tk.Entry(root, textvariable=column_num_var, width=50).grid(row=3, column=1, padx=10, pady=5)

# Function to open file dialog and set file path 
def open_file():
    open_name_var.set(filedialog.askopenfilename())

# Button to open file dialog
tk.Button(root, text="Browse", command=open_file).grid(row=0, column=2, padx=10, pady=5)

# Function to submit and close
def submit():
    global open_name, sheet_name, sheet_name_new, column_num
    open_name = open_name_var.get()
    sheet_name = sheet_name_var.get()
    sheet_name_new = sheet_name_new_var.get()
    column_num = column_num_var.get()
    root.destroy()

# Submit button
tk.Button(root, text="Submit", command=submit).grid(row=4, column=1, pady=10)

root.mainloop()

#Excel load
wb = load_workbook(open_name)
ws = wb[sheet_name]

#Get the header
header = []
for cell_row in ws[1]:
    header.append(cell_row.value)

# Dictionary to hold workbooks for each unique value in column 'F'
wb_dict = {}

# Iterate over the rows in the worksheet
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
    print(f"{row[0].row}.row from the total of {ws.max_row} rows")
    value = row[column_num-1].value  # Column 'F' (0-indexed as 5)

    if value not in wb_dict:
        # Create a new workbook for the unique value
        wb_dict[value] = Workbook()
        ws_new = wb_dict[value].active
        ws_new.title = sheet_name_new

        # Copy the header
        ws_new.append(header)
        for cell_orig, cell_new in zip(ws[1], ws_new[1]):
            cell_new.alignment = copy(cell_orig.alignment)
            cell_new.number_format = copy(cell_orig.number_format)
            cell_new.font = copy(cell_orig.font)
            cell_new.fill = copy(cell_orig.fill)
            cell_new.border = copy(cell_orig.border)
        
        for col_num in range(1,ws.max_column+1):
            ws_new.column_dimensions[get_column_letter(col_num)].width = ws.column_dimensions[get_column_letter(col_num)].width
            
    # Append the row to the appropriate workbook
    ws_new = wb_dict[value].active
    ws_new.append([cell.value for cell in row])

    # Copy styles
    for cell_orig, cell_new in zip(row, ws_new[ws_new.max_row]):
        cell_new.alignment = copy(cell_orig.alignment)
        cell_new.number_format = copy(cell_orig.number_format)
        cell_new.font = copy(cell_orig.font)
        cell_new.fill = copy(cell_orig.fill)
        cell_new.border = copy(cell_orig.border)

# Save all workbooks
save_name = filedialog.askdirectory()
excel_name = tk.simpledialog.askstring("Input", "What do you want to name the files?")
for value, wb in wb_dict.items():
    wb.save(save_name + f"/{value}_{excel_name}.xlsx")
root.destroy()

